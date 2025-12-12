import pandas as pd
from flask import send_file
from datetime import datetime
from io import BytesIO
from models.tuban import Tuban
from utils.helpers import parse_date

def import_tubans_from_excel(filepath):
    """从Excel文件导入图斑数据"""
    try:
        # 读取Excel文件
        df = pd.read_excel(filepath)

        # 标准化列名
        df.columns = df.columns.str.strip()

        # 必要的列映射
        column_mapping = {
            '图斑编号': 'tuban_code',
            '所属地质公园名称': 'park_name',
            '所在功能区': 'func_zone',
            '活动/设施名称': 'facility_name',
            '经度': 'longitude',
            '纬度': 'latitude',
            '占地面积': 'area',
            '影像时相': 'image_date',
            '建设单位': 'build_unit',
            '建设时间': 'build_time',
            '是否有审批手续': 'has_approval',
            '审批文号': 'approval_no',
            '发现时间': 'discover_time',
            '发现方式': 'discover_method',
            '现场核查时间': 'check_time',
            '核查人员': 'check_person',
            '核查结论': 'check_result',
            '问题类型': 'problem_type',
            '问题描述': 'problem_desc',
            '涉及地质遗迹类型': 'geo_heritage_type',
            '影响程度': 'impact_level',
            '是否违法违规': 'is_illegal',
            '违反法规条款': 'violated_law',
            '整改措施': 'rectify_measure',
            '整改时限': 'rectify_deadline',
            '整改进展': 'rectify_status',
            '整改验收时间': 'rectify_verify_time',
            '验收人员': 'verify_person',
            '是否销号': 'is_closed',
            '是否处罚': 'is_punished',
            '处罚形式': 'punish_type',
            '罚款金额': 'fine_amount',
            '处罚文书编号': 'punish_doc_no',
            '台账来源': 'data_source',
            '是否为巡查点': 'is_patrol_point',
            '责任部门/责任人': 'responsible_dept',
            '附件材料': 'attachments',
            '备注': 'remark'
        }

        # 重命名列
        df = df.rename(columns=column_mapping)

        # 只保留存在的列
        existing_columns = [col for col in column_mapping.values() if col in df.columns]
        df = df[existing_columns]

        # 处理日期列
        date_columns = ['image_date', 'build_time', 'discover_time', 'check_time', 'rectify_deadline', 'rectify_verify_time']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')

        # 处理数值列
        numeric_columns = ['longitude', 'latitude', 'area', 'fine_amount']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # 必填字段检查
        required_fields = ['tuban_code', 'park_name']
        df = df.dropna(subset=required_fields)

        # 导入数据
        count = 0
        for _, row in df.iterrows():
            # 检查是否已存在
            if Tuban.query.filter_by(tuban_code=row['tuban_code']).first():
                continue

            # 创建图斑对象
            tuban = Tuban()

            # 设置字段值
            for col in df.columns:
                if hasattr(tuban, col) and pd.notna(row[col]):
                    setattr(tuban, col, row[col])

            # 设置默认值
            if not tuban.rectify_status:
                tuban.rectify_status = '未整改'
            if not tuban.is_closed:
                tuban.is_closed = '否'
            if not tuban.is_punished:
                tuban.is_punished = '否'
            if not tuban.is_patrol_point:
                tuban.is_patrol_point = '否'
            if not tuban.has_approval:
                tuban.has_approval = '否'
            if not tuban.is_illegal:
                tuban.is_illegal = '待定'

            # 保存到数据库
            from models import db
            db.session.add(tuban)
            count += 1

        db.session.commit()
        return count

    except Exception as e:
        from models import db
        db.session.rollback()
        raise Exception(f"Excel导入失败: {str(e)}")

def export_tubans_to_excel(tubans):
    """导出图斑数据到Excel"""
    try:
        # 准备数据
        data = []
        for tuban in tubans:
            data.append({
                '图斑编号': tuban.tuban_code,
                '所属地质公园名称': tuban.park_name,
                '所在功能区': tuban.func_zone,
                '活动/设施名称': tuban.facility_name,
                '经度': tuban.longitude,
                '纬度': tuban.latitude,
                '占地面积': tuban.area,
                '影像时相': tuban.image_date,
                '建设单位': tuban.build_unit,
                '建设时间': tuban.build_time,
                '是否有审批手续': tuban.has_approval,
                '审批文号': tuban.approval_no,
                '发现时间': tuban.discover_time,
                '发现方式': tuban.discover_method,
                '现场核查时间': tuban.check_time,
                '核查人员': tuban.check_person,
                '核查结论': tuban.check_result,
                '问题类型': tuban.problem_type,
                '问题描述': tuban.problem_desc,
                '涉及地质遗迹类型': tuban.geo_heritage_type,
                '影响程度': tuban.impact_level,
                '是否违法违规': tuban.is_illegal,
                '违反法规条款': tuban.violated_law,
                '整改措施': tuban.rectify_measure,
                '整改时限': tuban.rectify_deadline,
                '整改进展': tuban.rectify_status,
                '整改验收时间': tuban.rectify_verify_time,
                '验收人员': tuban.verify_person,
                '是否销号': tuban.is_closed,
                '是否处罚': tuban.is_punished,
                '处罚形式': tuban.punish_type,
                '罚款金额': tuban.fine_amount,
                '处罚文书编号': tuban.punish_doc_no,
                '台账来源': tuban.data_source,
                '是否为巡查点': tuban.is_patrol_point,
                '责任部门/责任人': tuban.responsible_dept,
                '附件材料': tuban.attachments,
                '备注': tuban.remark
            })

        # 创建DataFrame
        df = pd.DataFrame(data)

        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='图斑数据', index=False)

            # 获取工作表
            worksheet = writer.sheets['图斑数据']

            # 设置列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 设置标题行样式
            from openpyxl.styles import Font, Alignment
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

        # 准备下载
        output.seek(0)
        filename = f"图斑数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            BytesIO(output.getvalue()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        raise Exception(f"Excel导出失败: {str(e)}")

def create_excel_template():
    """创建Excel导入模板"""
    try:
        # 定义模板列
        columns = [
            '图斑编号', '所属地质公园名称', '所在功能区', '活动/设施名称', '经度', '纬度',
            '占地面积', '影像时相', '建设单位', '建设时间', '是否有审批手续', '审批文号',
            '发现时间', '发现方式', '现场核查时间', '核查人员', '核查结论', '问题类型',
            '问题描述', '涉及地质遗迹类型', '影响程度', '是否违法违规', '违反法规条款',
            '整改措施', '整改时限', '整改进展', '整改验收时间', '验收人员', '是否销号',
            '是否处罚', '处罚形式', '罚款金额', '处罚文书编号', '台账来源', '是否为巡查点',
            '责任部门/责任人', '附件材料', '备注'
        ]

        # 创建示例数据
        sample_data = [
            ['TB001', '张家界世界地质公园', '核心区', '违规观景台', 110.4833, 29.3167, 500, '2023-05-15',
             '某旅游公司', '2022-08-01', '否', '', '2023-06-01', '遥感监测', '2023-06-10', '张三',
             '经核查为违规建设', '违规建设', '破坏自然景观', '地层剖面', '严重', '是', '《地质遗迹保护管理规定》',
             '立即拆除', '2023-12-31', '整改中', '', '', '否', '否', '', '', '', '遥感监测台账',
             '是', '地质公园管理处', '', ''],
            ['TB002', '嵩山世界地质公园', '缓冲区', '非法采矿', 112.9531, 34.4864, 1200, '2023-07-20',
             '某矿业公司', '2023-03-01', '否', '', '2023-08-05', '日常巡查', '2023-08-08', '李四',
             '确认为非法采矿', '采矿', '破坏地质剖面', '构造遗迹', '严重', '是', '《矿产资源法》',
             '封堵矿口，恢复植被', '2023-10-31', '已整改', '2023-10-25', '王五', '是', '是', '罚款',
             200000, '豫矿罚〔2023〕15号', '巡查台账', '是', '国土资源执法大队', '', '']
        ]

        # 创建DataFrame
        df = pd.DataFrame(sample_data, columns=columns)

        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='模板', index=False)

            # 获取工作表
            worksheet = writer.sheets['模板']

            # 设置样式
            from openpyxl.styles import Font, Alignment, PatternFill

            # 标题行样式
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # 设置列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 添加说明工作表
            instructions = [
                ['填写说明'],
                ['1. 红色标题为必填字段'],
                ['2. 日期格式：YYYY-MM-DD'],
                ['3. 坐标请使用十进制度格式'],
                ['4. 面积单位为平方米'],
                ['5. 罚款金额单位为万元'],
                ['6. 请勿修改工作表名称'],
                [''],
                ['字典选项（请从以下选项中选择）'],
                ['功能区：核心区、缓冲区、实验区'],
                ['发现方式：遥感监测、日常巡查、群众举报、上级交办'],
                ['问题类型：违规建设、采矿、开垦、污染'],
                ['影响程度：严重、一般、轻微'],
                ['整改进展：未整改、整改中、已整改'],
                ['是否选项：是、否'],
                ['违法违规选项：是、否、待定']
            ]

            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='填写说明', index=False, header=False)

            # 设置说明工作表样式
            instruction_sheet = writer.sheets['填写说明']
            instruction_sheet.column_dimensions['A'].width = 60
            for row in instruction_sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='center')

        output.seek(0)
        filename = f"图斑导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            BytesIO(output.getvalue()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        raise Exception(f"创建模板失败: {str(e)}")